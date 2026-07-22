import os
import io
import uuid
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

from flask import current_app
from werkzeug.utils import secure_filename

ALLOWED_IMAGE_EXTENSIONS = {"png", "jpg", "jpeg", "webp"}
ALLOWED_RECEIPT_EXTENSIONS = ALLOWED_IMAGE_EXTENSIONS | {"pdf"}


def send_email(to_email, subject, body):
    """
    Send a plain-text email via SMTP (configured for Gmail by default).
    Never raises — returns True on success, False otherwise, so a failed/unconfigured
    email never breaks the request that triggered it (approval, guarantor notice, etc).
    """
    if not to_email:
        return False

    mail_server = os.environ.get("MAIL_SERVER", "smtp.gmail.com")
    mail_port = int(os.environ.get("MAIL_PORT", 587))
    mail_username = os.environ.get("MAIL_USERNAME")
    mail_password = os.environ.get("MAIL_PASSWORD")
    mail_sender = os.environ.get("MAIL_SENDER", mail_username)

    if not mail_username or not mail_password:
        print(f"[email] Skipped sending to {to_email}: MAIL_USERNAME/MAIL_PASSWORD not configured.")
        return False

    msg = MIMEMultipart()
    msg["From"] = mail_sender
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    try:
        with smtplib.SMTP(mail_server, mail_port, timeout=10) as server:
            server.starttls()
            server.login(mail_username, mail_password)
            server.sendmail(mail_sender, [to_email], msg.as_string())
        return True
    except Exception as e:
        print(f"[email] Failed to send to {to_email}: {e}")
        return False


def allowed_image(filename, allowed_extensions=None):
    allowed_extensions = allowed_extensions or ALLOWED_IMAGE_EXTENSIONS
    return "." in filename and filename.rsplit(".", 1)[1].lower() in allowed_extensions


def _max_upload_bytes():
    mb = float(current_app.config.get("MAX_UPLOAD_MB", 5))
    return int(mb * 1024 * 1024)


_CONTENT_TYPES = {
    "png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg",
    "webp": "image/webp", "pdf": "application/pdf",
}

_supabase_client_singleton = None


def _supabase_client():
    """
    Lazily build (and cache per-process) the Supabase client used for file storage.
    Uses the service role key so the server can read/write regardless of bucket
    policies — the bucket itself should be PRIVATE; access to files is gated by
    our own @login_required / admin-only routes, not by public Supabase URLs.
    """
    global _supabase_client_singleton
    if _supabase_client_singleton is None:
        from supabase import create_client
        url = os.environ.get("SUPABASE_URL")
        key = os.environ.get("SUPABASE_SERVICE_KEY")
        if not url or not key:
            raise RuntimeError(
                "SUPABASE_URL and SUPABASE_SERVICE_KEY must be set in the environment "
                "to upload, read, or delete member files."
            )
        _supabase_client_singleton = create_client(url, key)
    return _supabase_client_singleton


def _supabase_bucket():
    return os.environ.get("SUPABASE_BUCKET", "cms-uploads")


def save_upload(file_storage, subfolder, allowed_extensions=None):
    """
    Upload a file to Supabase Storage, at <subfolder>/<uuid>_<filename>.
    Returns the relative path stored in the DB (subfolder/filename), or None.
    Raises ValueError on invalid file / oversize / upload failure.
    """
    if not file_storage or file_storage.filename == "":
        return None

    filename = secure_filename(file_storage.filename)
    if not allowed_image(filename, allowed_extensions):
        allowed = allowed_extensions or ALLOWED_IMAGE_EXTENSIONS
        raise ValueError(f"Only {'/'.join(sorted(e.upper() for e in allowed))} files are allowed.")

    file_storage.seek(0, os.SEEK_END)
    size = file_storage.tell()
    file_storage.seek(0)
    if size > _max_upload_bytes():
        raise ValueError(
            f"File too large. Max allowed is {current_app.config.get('MAX_UPLOAD_MB', 5)}MB."
        )

    unique_name = f"{uuid.uuid4().hex}_{filename}"
    relative_path = f"{subfolder}/{unique_name}"
    file_bytes = file_storage.read()
    ext = filename.rsplit(".", 1)[-1].lower()
    content_type = _CONTENT_TYPES.get(ext, "application/octet-stream")

    try:
        _supabase_client().storage.from_(_supabase_bucket()).upload(
            relative_path, file_bytes, {"content-type": content_type}
        )
    except Exception as e:
        raise ValueError(f"Upload to storage failed: {e}")

    return relative_path


def get_upload_bytes(relative_path):
    """
    Download a previously uploaded file's raw bytes from Supabase Storage.
    Returns None if the path is empty, or the file can't be found/downloaded
    (e.g. an old record from before the Supabase migration, or a transient error).
    """
    if not relative_path:
        return None
    # Normalize in case older DB records were saved with a Windows-style backslash
    # (os.path.join on Windows produces "signatures\file.jpg" rather than "signatures/file.jpg").
    normalized = relative_path.replace("\\", "/")
    try:
        return _supabase_client().storage.from_(_supabase_bucket()).download(normalized)
    except Exception:
        return None


def delete_upload(relative_path):
    if not relative_path:
        return
    normalized = relative_path.replace("\\", "/")
    try:
        _supabase_client().storage.from_(_supabase_bucket()).remove([normalized])
    except Exception:
        pass


def _wrap_text(text, font_name, font_size, max_width, canvas_obj):
    """Simple word-wrap helper: splits text into lines that fit max_width."""
    words = text.split()
    lines, current = [], ""
    for word in words:
        candidate = f"{current} {word}".strip()
        if canvas_obj.stringWidth(candidate, font_name, font_size) <= max_width:
            current = candidate
        else:
            if current:
                lines.append(current)
            current = word
    if current:
        lines.append(current)
    return lines


def build_member_profile_pdf(user):
    """
    Generate a filled copy of the cooperative's official Membership Application Form
    (matching MOLETE (IB) BEST CHOICE MULTIPURPOSE COOPERATIVE SOCIETY LTD's paper form),
    populated with whatever data the system has on file. Fields the system doesn't collect
    (marital status, sex, employer, next of kin, bankers/branch, starting contribution, etc.)
    are left as blank lines, same as the original form, for manual completion.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    margin = 18 * mm
    content_width = width - 2 * margin

    def letterhead(y_top):
        c.setFont("Helvetica-Bold", 13)
        c.drawCentredString(width / 2, y_top, "MOLETE (IB) BEST CHOICE MULTIPURPOSE COOPERATIVE SOCIETY LTD")
        c.setFont("Helvetica-Bold", 9)
        c.drawCentredString(width / 2, y_top - 6 * mm, "C/O FBN Building 48 Molete/Challenge, Opposite Challenge R/About, Ibadan Oyo State")
        c.setFont("Helvetica", 9)
        c.drawCentredString(width / 2, y_top - 11 * mm, "E-mail: bestchoicecooperative@gmail.com")
        c.setLineWidth(0.7)
        c.line(margin, y_top - 15 * mm, width - margin, y_top - 15 * mm)
        return y_top - 15 * mm

    def field_line(label, value, x, y, label_width=42 * mm, line_width=None):
        c.setFont("Helvetica-Bold", 9.5)
        c.drawString(x, y, label)
        c.setFont("Helvetica", 9.5)
        lw = line_width if line_width else (content_width - label_width - (x - margin))
        line_x = x + label_width
        if value:
            c.drawString(line_x + 1 * mm, y, str(value))
        c.setLineWidth(0.5)
        c.line(line_x, y - 1.2 * mm, line_x + lw, y - 1.2 * mm)

    # ================= PAGE 1: Membership Application Form =================
    y = height - 15 * mm
    y = letterhead(y)

    # Passport photo box, top-right — given its own clear vertical band. All text below (title
    # included) is positioned relative to the box's bottom edge, so nothing can ever run under it.
    passport_w, passport_h = 30 * mm, 35 * mm
    passport_x = width - margin - passport_w
    passport_top = y - 4 * mm
    passport_y = passport_top - passport_h
    c.setFont("Helvetica", 7)
    c.rect(passport_x, passport_y, passport_w, passport_h)
    if user.passport_path:
        try:
            from reportlab.lib.utils import ImageReader
            passport_bytes = get_upload_bytes(user.passport_path)
            c.drawImage(ImageReader(io.BytesIO(passport_bytes)), passport_x, passport_y,
                        width=passport_w, height=passport_h, preserveAspectRatio=True, anchor='c')
        except Exception:
            c.drawCentredString(passport_x + passport_w / 2, passport_y + passport_h / 2, "Affix Passport")
    else:
        c.drawCentredString(passport_x + passport_w / 2, passport_y + passport_h / 2, "Affix Passport")

    # Title sits beside the box (left-aligned, well clear of the box's left edge — not centered,
    # since a centered title's right half would otherwise run under the box).
    c.setFont("Helvetica-Bold", 14)
    title_y = passport_top - passport_h / 2 + 2 * mm
    c.drawString(margin, title_y, "MEMBERSHIP APPLICATION FORM")

    # Everything from here on starts strictly below the box's bottom edge — guaranteed no overlap.
    y = passport_y - 8 * mm
    c.setFont("Helvetica-Bold", 10.5)
    c.drawString(margin, y, "A)  To be completed by applicant")
    y -= 7 * mm

    reg_date = user.created_at or __import__("datetime").datetime.utcnow()
    intro = (
        f"I, {user.full_name} of ______________________________________________________ "
        f"on this {reg_date.strftime('%d')} day of {reg_date.strftime('%B')} {reg_date.strftime('%Y')} "
        f"hereby apply for admission as a member of Best Choice Multipurpose Cooperative Society "
        f"Limited (or any other name she may be known in future) and remit herewith an application "
        f"& Development fee of N"
        f"{'{:,.2f}'.format(user.application_fee) if user.application_fee is not None else '______________'}."
    )
    c.setFont("Helvetica", 9.5)
    for line in _wrap_text(intro, "Helvetica", 9.5, content_width, c):
        c.drawString(margin, y, line)
        y -= 5 * mm

    y -= 1 * mm
    liability = (
        "On my admission as a member by the society, I shall be jointly and severally liable to "
        "discharge the present and future liabilities of the society, and shall respect the laws "
        "guiding the society."
    )
    for line in _wrap_text(liability, "Helvetica", 9.5, content_width, c):
        c.drawString(margin, y, line)
        y -= 5 * mm

    y -= 4 * mm
    field_line("Marital Status:", user.marital_status, margin, y, label_width=28 * mm, line_width=28 * mm)
    field_line("Sex:", user.sex, margin + 65 * mm, y, label_width=12 * mm, line_width=20 * mm)
    field_line("Profession:", user.work_position, margin + 105 * mm, y, label_width=22 * mm, line_width=content_width - 127 * mm)
    y -= 8 * mm
    field_line("Email:", user.email, margin, y, label_width=16 * mm)
    y -= 8 * mm
    field_line("Employer:", user.employer, margin, y, label_width=20 * mm)
    y -= 8 * mm
    field_line("Department/Section:", user.work_position, margin, y, label_width=34 * mm)
    y -= 8 * mm
    field_line("Tel.:", user.phone_number, margin, y, label_width=12 * mm, line_width=50 * mm)
    field_line("Office No.:", user.office_number, margin + 68 * mm, y, label_width=22 * mm, line_width=content_width - 90 * mm)
    y -= 8 * mm
    field_line("Bankers/Branch:", user.bankers_branch, margin, y, label_width=30 * mm, line_width=50 * mm)
    field_line("A/C No.:", user.account_number, margin + 85 * mm, y, label_width=18 * mm, line_width=content_width - 103 * mm)
    y -= 8 * mm
    starting_contrib = f"{user.starting_contribution:,.2f}" if user.starting_contribution is not None else ""
    field_line("Starting Contribution (N):", starting_contrib, margin, y, label_width=44 * mm, line_width=35 * mm)
    field_line("Deduction Due Date:", user.deduction_due_date, margin + 85 * mm, y, label_width=32 * mm, line_width=content_width - 117 * mm)
    y -= 8 * mm
    field_line("Preferred Month of Deduction and year:", user.preferred_deduction_month_year, margin, y, label_width=62 * mm)

    y -= 6 * mm
    c.setFont("Helvetica-Oblique", 8.5)
    c.drawString(margin, y, "Molete (Ibadan) Best Choice Multipurpose cooperative Society account number is 3059522679")

    # Section B - Next of Kin
    y -= 10 * mm
    c.setFont("Helvetica-Bold", 10.5)
    c.drawString(margin, y, "B)  Next of Kin")
    y -= 7 * mm
    field_line("Name:", user.next_of_kin_name, margin, y, label_width=16 * mm)
    y -= 8 * mm
    field_line("Address:", user.next_of_kin_address, margin, y, label_width=20 * mm)
    y -= 8 * mm
    field_line("Tel:", user.next_of_kin_phone, margin, y, label_width=12 * mm, line_width=60 * mm)

    # Applicant signature
    y -= 16 * mm
    sig_w, sig_h = 55 * mm, 18 * mm
    if user.signature_path:
        try:
            from reportlab.lib.utils import ImageReader
            sig_bytes = get_upload_bytes(user.signature_path)
            c.drawImage(ImageReader(io.BytesIO(sig_bytes)), margin, y, width=sig_w, height=sig_h,
                        preserveAspectRatio=True, anchor='c')
        except Exception:
            pass
    c.setLineWidth(0.5)
    c.line(margin, y - 1 * mm, margin + sig_w, y - 1 * mm)
    c.setFont("Helvetica-Bold", 9)
    c.drawString(margin, y - 6 * mm, "APPLICANT SIGNATURE")

    c.setFont("Helvetica-Oblique", 7.5)
    c.drawString(margin, 12 * mm, f"System-generated on behalf of {user.full_name} ({user.membership_code or 'pending'}) — confidential.")

    # ================= PAGE 2: Membership Certificate =================
    c.showPage()
    y = height - 15 * mm
    y = letterhead(y)
    y -= 14 * mm

    c.setFont("Helvetica-Bold", 10.5)
    c.drawString(margin, y, "C)  Membership Certificate")
    y -= 10 * mm

    today = __import__("datetime").datetime.utcnow()
    cert_text = (
        f"This certifies that {user.full_name} of {user.work_position or '______________________'} "
        f"is a member of Best Choice Multipurpose Cooperative Society Limited and is entitled to all "
        f"of the rights, benefits, and privileges of the Society."
    )
    c.setFont("Helvetica", 10.5)
    for line in _wrap_text(cert_text, "Helvetica", 10.5, content_width, c):
        c.drawString(margin, y, line)
        y -= 6 * mm

    y -= 4 * mm
    field_line("Date:", today.strftime("%d %B %Y"), margin, y, label_width=16 * mm, line_width=50 * mm)

    y -= 30 * mm
    c.setLineWidth(0.5)
    c.line(margin, y, margin + 60 * mm, y)
    c.setFont("Helvetica", 9.5)
    c.drawString(margin, y - 5 * mm, "(President)")

    c.line(margin + 90 * mm, y, margin + 150 * mm, y)
    c.drawString(margin + 90 * mm, y - 5 * mm, "(Secretary)")

    c.setFont("Helvetica-Oblique", 7.5)
    c.drawString(margin, 12 * mm, f"System-generated on behalf of {user.full_name} ({user.membership_code or 'pending'}) — confidential.")

    # ================= PAGE 3: Mandate Instruction =================
    c.showPage()
    y = height - 15 * mm
    y = letterhead(y)
    y -= 12 * mm

    c.setFont("Helvetica-Bold", 12)
    c.drawString(margin, y, "MANDATE INSTRUCTION")
    y -= 8 * mm

    c.setFont("Helvetica", 9.5)
    intro_lines = [
        f"From: {user.full_name}",
        f"Name: {user.full_name}",
        f"Staff no.: {user.account_number or '________________'}",
        "To: Molete (Ibadan) Best Choice Multipurpose Cooperative Society Limited",
        "",
        "Kindly take this as an instruction to debit my salary account number __________________________ on a monthly basis as itemized below:",
        "",
        "Contribution: N __________________",
        "Ordinary Share Capital (if any): N __________________",
        "Investment Share Capital (if any): N __________________",
        "Development levy: N __________________",
        "",
        "In any event where I obtain/have obtained a loan(s) from the cooperative; kindly deduct the monthly repayment of such loan(s) from my salary account in addition to deductions itemized above. The cooperative has the right to spread my monthly loan repayment default over the remainder months for the loan repayment tenor (inclusive of default fee if any) and such should be deducted from my salary account.",
        "",
        "You also have my concurrence to lien my salary account stated above to the tune of my monthly loan repayment (inclusive of default fee if any) once my salary account is not funded for the month’s deduction.",
        "",
        "However, in an event where I cease (which shall be communicated to you formally) to be a member of the cooperative, this standing instruction becomes ineffective while the balance of any loan I obtained from the cooperative should be deducted from my salary account. Molete (Ibadan) Best Choice Multipurpose Cooperative Society Limited has the right to inform my employer should I fail to pay the debt owed.",
        "",
        "SIGNED",
        "____________________________",
    ]
    for line in intro_lines:
        if not line:
            y -= 4 * mm
            continue
        for wrapped in _wrap_text(line, "Helvetica", 9.5, content_width, c):
            c.drawString(margin, y, wrapped)
            y -= 4.2 * mm

    y -= 8 * mm
    sig_w, sig_h = 55 * mm, 18 * mm
    if user.signature_path:
        try:
            from reportlab.lib.utils import ImageReader
            sig_bytes = get_upload_bytes(user.signature_path)
            c.drawImage(ImageReader(io.BytesIO(sig_bytes)), margin, y, width=sig_w, height=sig_h,
                        preserveAspectRatio=True, anchor='c')
        except Exception:
            pass
    c.setLineWidth(0.5)
    c.line(margin, y - 1 * mm, margin + sig_w, y - 1 * mm)

    c.setFont("Helvetica-Oblique", 7.5)
    c.drawString(margin, 12 * mm, f"System-generated on behalf of {user.full_name} ({user.membership_code or 'pending'}) — confidential.")

    c.showPage()
    c.save()
    buf.seek(0)
    return buf


def _brief_letterhead(c, width, y_top, margin):
    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(width / 2, y_top, "MOLETE (IB) BEST CHOICE MULTIPURPOSE COOPERATIVE SOCIETY LTD")
    from reportlab.lib.units import mm
    c.setFont("Helvetica", 8.5)
    c.drawCentredString(width / 2, y_top - 5.5 * mm, "Member Signature Brief")
    c.setLineWidth(0.7)
    c.line(margin, y_top - 9 * mm, width - margin, y_top - 9 * mm)
    return y_top - 9 * mm


def build_member_brief_pdf(user):
    """
    Compact single-page brief for one member: name, membership code, account number,
    and their signature image. Intended as a lightweight alternative to embedding
    raw signature file paths/images in the Excel export.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    margin = 20 * mm

    y = height - 20 * mm
    y = _brief_letterhead(c, width, y, margin)
    y -= 16 * mm

    def field(label, value, y):
        c.setFont("Helvetica-Bold", 11)
        c.drawString(margin, y, label)
        c.setFont("Helvetica", 11)
        c.drawString(margin + 45 * mm, y, str(value) if value else "-")
        c.setLineWidth(0.4)
        c.line(margin + 45 * mm, y - 1.5 * mm, width - margin, y - 1.5 * mm)

    field("Full Name:", user.full_name, y)
    y -= 12 * mm
    field("Membership Code:", user.membership_code or "-", y)
    y -= 12 * mm
    field("Account Number:", user.account_number, y)
    y -= 20 * mm

    c.setFont("Helvetica-Bold", 11)
    c.drawString(margin, y, "Signature:")
    sig_box_w, sig_box_h = 70 * mm, 25 * mm
    sig_box_y = y - sig_box_h - 4 * mm
    c.setLineWidth(0.6)
    c.rect(margin, sig_box_y, sig_box_w, sig_box_h)
    if user.signature_path:
        sig_bytes = get_upload_bytes(user.signature_path)
        if sig_bytes:
            try:
                from reportlab.lib.utils import ImageReader
                c.drawImage(ImageReader(io.BytesIO(sig_bytes)), margin + 2 * mm, sig_box_y + 2 * mm,
                            width=sig_box_w - 4 * mm, height=sig_box_h - 4 * mm,
                            preserveAspectRatio=True, anchor='c', mask='auto')
            except Exception:
                c.setFont("Helvetica-Oblique", 8)
                c.drawCentredString(margin + sig_box_w / 2, sig_box_y + sig_box_h / 2, "Signature unavailable")
        else:
            c.setFont("Helvetica-Oblique", 8)
            c.drawCentredString(margin + sig_box_w / 2, sig_box_y + sig_box_h / 2, "Signature unavailable")
    else:
        c.setFont("Helvetica-Oblique", 8)
        c.drawCentredString(margin + sig_box_w / 2, sig_box_y + sig_box_h / 2, "No signature on file")

    c.setFont("Helvetica-Oblique", 7.5)
    c.drawString(margin, 12 * mm, f"System-generated brief for {user.full_name} ({user.membership_code or 'pending'}) — confidential.")

    c.showPage()
    c.save()
    buf.seek(0)
    return buf


def build_members_brief_pdf_combined(users):
    """
    Single PDF listing every member's name, membership code, account number and
    signature image, several members per page.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    margin = 15 * mm
    row_h = 28 * mm
    sig_w, sig_h = 45 * mm, 18 * mm

    y = height - 15 * mm
    y = _brief_letterhead(c, width, y, margin)
    y -= 8 * mm

    def header_row(y):
        c.setFont("Helvetica-Bold", 8.5)
        c.drawString(margin, y, "Name")
        c.drawString(margin + 55 * mm, y, "Membership Code")
        c.drawString(margin + 95 * mm, y, "Account No.")
        c.drawString(margin + 130 * mm, y, "Signature")
        c.setLineWidth(0.5)
        c.line(margin, y - 2 * mm, width - margin, y - 2 * mm)
        return y - 6 * mm

    y = header_row(y)

    for u in users:
        if y - row_h < 20 * mm:
            c.showPage()
            y = height - 15 * mm
            y = _brief_letterhead(c, width, y, margin)
            y -= 8 * mm
            y = header_row(y)

        text_y = y - row_h / 2 + 2 * mm
        c.setFont("Helvetica", 9)
        c.drawString(margin, text_y, (u.full_name or "")[:32])
        c.drawString(margin + 55 * mm, text_y, u.membership_code or "-")
        c.drawString(margin + 95 * mm, text_y, u.account_number or "-")

        sig_x = margin + 130 * mm
        sig_y = y - row_h + (row_h - sig_h) / 2
        c.setLineWidth(0.4)
        c.rect(sig_x, sig_y, sig_w, sig_h)
        if u.signature_path:
            sig_bytes = get_upload_bytes(u.signature_path)
            if sig_bytes:
                try:
                    from reportlab.lib.utils import ImageReader
                    c.drawImage(ImageReader(io.BytesIO(sig_bytes)), sig_x + 1 * mm, sig_y + 1 * mm,
                                width=sig_w - 2 * mm, height=sig_h - 2 * mm,
                                preserveAspectRatio=True, anchor='c', mask='auto')
                except Exception:
                    pass

        c.setLineWidth(0.3)
        c.line(margin, y - row_h, width - margin, y - row_h)
        y -= row_h

    c.showPage()
    c.save()
    buf.seek(0)
    return buf


def build_members_csv_minimal(users):
    """CSV export with the core member fields plus signature metadata."""
    import csv

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Full Name", "Membership Code", "Account Number", "Office Number", "Signature Present", "Signature File"])
    for u in users:
        writer.writerow([
            u.full_name,
            u.membership_code or "",
            u.account_number,
            u.office_number or "",
            "Yes" if u.signature_path else "No",
            u.signature_path or "",
        ])

    byte_buf = io.BytesIO(buf.getvalue().encode("utf-8"))
    byte_buf.seek(0)
    return byte_buf


def build_members_excel(users, brief_pdf_url_func=None):
    """
    brief_pdf_url_func: optional callable(user) -> absolute URL for that member's
    individual Brief PDF, used to populate the 'Brief PDF' column as a clickable link.
    """
    import pandas as pd
    from openpyxl import load_workbook

    data = [{
        "Membership Code": u.membership_code,
        "Full Name": u.full_name,
        "Account Number": u.account_number,
        "Phone Number": u.phone_number,
        "Office Number": u.office_number,
        "Email": u.email,
        "Work Position": u.work_position,
        "Employer": u.employer,
        "Marital Status": u.marital_status,
        "Sex": u.sex,
        "Bankers/Branch": u.bankers_branch,
        "Application Fee": u.application_fee,
        "Starting Contribution": u.starting_contribution,
        "Deduction Due Date": u.deduction_due_date,
        "Preferred Deduction Month/Year": u.preferred_deduction_month_year,
        "Next of Kin Name": u.next_of_kin_name,
        "Next of Kin Address": u.next_of_kin_address,
        "Next of Kin Phone": u.next_of_kin_phone,
        "Signature Present": "Yes" if u.signature_path else "No",
        "Brief PDF": "Open Brief PDF" if u.signature_path else "",
        "Savings Balance": u.savings_balance,
        "Status": u.account_status,
        "Profile Complete": "Yes" if u.is_profile_complete else "No",
    } for u in users]
    df = pd.DataFrame(data)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Members")

    workbook = load_workbook(buf)
    sheet = workbook["Members"]

    # Find the "Brief PDF" column letter dynamically (kept robust to column reordering).
    header_row = next(sheet.iter_rows(min_row=1, max_row=1))
    brief_col_idx = None
    for cell in header_row:
        if cell.value == "Brief PDF":
            brief_col_idx = cell.column_letter
            break

    if brief_col_idx and brief_pdf_url_func:
        for idx, u in enumerate(users, start=2):
            if not u.signature_path:
                continue
            cell = sheet[f"{brief_col_idx}{idx}"]
            cell.value = "Open Brief PDF"
            cell.hyperlink = brief_pdf_url_func(u)
            cell.style = "Hyperlink"

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def build_gift_preferences_excel(rows):
    """rows: list of (User, GiftPreference)"""
    import pandas as pd

    data = [{
        "Membership Code": u.membership_code,
        "Full Name": u.full_name,
        "Account Number": u.account_number,
        "Preference": gp.preference_type,
    } for u, gp in rows]
    df = pd.DataFrame(data)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Gift Preferences")
    buf.seek(0)
    return buf


def parse_bulk_import_file(file_storage):
    """
    Parse an uploaded CSV/Excel of existing members.
    Required columns (case-insensitive, common variants accepted): full_name, membership_code,
    account_number. Optional: work_position, phone_number.
    Email and other details are intentionally NOT collected here — members fill those in
    themselves when they log in for the first time and complete their profile.
    Returns a list of dicts.
    """
    import pandas as pd

    filename = file_storage.filename.lower()
    # dtype=str is essential here: without it, pandas infers account numbers and
    # membership codes as numeric and silently strips leading zeros (e.g. "08055556666" -> 8055556666).
    if filename.endswith(".csv"):
        df = pd.read_csv(file_storage, dtype=str, keep_default_na=False)
    elif filename.endswith(".xlsx") or filename.endswith(".xls"):
        df = pd.read_excel(file_storage, dtype=str)
        df = df.fillna("")
    else:
        raise ValueError("Please upload a .csv or .xlsx file.")

    df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]

    # Accept common real-world header variants without forcing the admin to rename columns.
    column_aliases = {
        "name": "full_name",
        "member_name": "full_name",
        "full_names": "full_name",
        "code": "membership_code",
        "member_code": "membership_code",
        "acct_number": "account_number",
        "acc_number": "account_number",
        "account_no": "account_number",
        "phone_no": "phone_number",
        "phone": "phone_number",
        "telephone": "phone_number",
        "department": "work_position",
        "position": "work_position",
    }
    df = df.rename(columns={k: v for k, v in column_aliases.items() if k in df.columns})

    required = {"full_name", "membership_code", "account_number"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(sorted(missing))}")

    records = []
    for _, row in df.iterrows():
        records.append({
            "full_name": str(row["full_name"]).strip(),
            "membership_code": str(row["membership_code"]).strip(),
            "account_number": str(row["account_number"]).strip(),
            "work_position": str(row.get("work_position", "")).strip() if "work_position" in df.columns else "",
            "phone_number": str(row.get("phone_number", "")).strip() if "phone_number" in df.columns else "",
        })
    return records
